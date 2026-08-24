"""Two-sheet (Nodes, Edges) .xlsx workbook: writer for export/template,
reader for import.

Columns are schema-driven — one per ``slot_names()`` entry, never a
hardcoded CAMO slot name. A slot whose value is a nested class (a LinkML
fieldset, e.g. CAMO's ``mediation``) doesn't fit a single scalar cell; its
column holds a compact JSON string instead, which
``graph_import._flatten_row()`` expands back into the "__"-joined key
scheme ``bind_form_data()`` expects on the way back in.
"""

from __future__ import annotations

import io
import json

import openpyxl

from .graph_import import ImportParseError

NODE_ID_COLUMNS = ["node_id", "name"]
EDGE_ID_COLUMNS = ["edge_id", "subject_node_id", "object_node_id", "predicate"]


def write_graph_xlsx(graph, *, blank: bool = False) -> bytes:
    """Two-sheet workbook. Columns from slot_names(); data rows from
    serialize_graph(graph) unless blank=True (headers only, i.e. the
    importable template)."""
    from apps.export.serializer import serialize_graph
    from apps.schemas.schema_engine import get_schema_view

    schema_view = get_schema_view(graph.schema_version, graph.document.project)
    node_slots = [
        s for s in schema_view.slot_names("CausalNode") if s not in NODE_ID_COLUMNS
    ]
    edge_slots = [
        s
        for s in schema_view.slot_names("CausalEdge")
        if s not in ("edge_id", "subject", "object", "predicate")
    ]

    data = None if blank else serialize_graph(graph)

    workbook = openpyxl.Workbook()
    nodes_ws = workbook.active
    nodes_ws.title = "Nodes"
    node_header = NODE_ID_COLUMNS + node_slots
    nodes_ws.append(node_header)
    if data:
        for node in data["nodes"]:
            nodes_ws.append([_cell_value(node.get(col)) for col in node_header])

    edges_ws = workbook.create_sheet("Edges")
    edge_header = EDGE_ID_COLUMNS + edge_slots
    edges_ws.append(edge_header)
    if data:
        for edge in data["edges"]:
            row = []
            for col in edge_header:
                if col == "subject_node_id":
                    row.append(edge.get("subject", ""))
                elif col == "object_node_id":
                    row.append(edge.get("object", ""))
                else:
                    row.append(_cell_value(edge.get(col)))
            edges_ws.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if value is None:
        return ""
    return value


def parse_xlsx_import(file_obj) -> tuple[list[dict], list[dict]]:
    """Read Nodes/Edges sheets -> (node_rows, edge_rows), each a
    header->cell dict. Blank rows are skipped."""
    try:
        workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises several exception types for bad files
        raise ImportParseError(f"Could not read the Excel file: {exc}") from exc

    if "Nodes" not in workbook.sheetnames or "Edges" not in workbook.sheetnames:
        raise ImportParseError("Workbook must contain 'Nodes' and 'Edges' sheets.")

    return _read_sheet(workbook["Nodes"]), _read_sheet(workbook["Edges"])


def _read_sheet(worksheet) -> list[dict]:
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return []

    out: list[dict] = []
    for raw_row in rows:
        if raw_row is None or all(cell in (None, "") for cell in raw_row):
            continue
        row = {col: _parse_cell(value) for col, value in zip(header, raw_row) if col}
        out.append(row)
    return out


def _parse_cell(value):
    if isinstance(value, str) and value[:1] in "{[":
        try:
            return json.loads(value)
        except (json.JSONDecodeError, ValueError):
            return value
    return value
